import openpyxl
import os.path
import requests
from tqdm import tqdm
import re
from requests_ntlm import HttpNtlmAuth

class DownloadSharepoint():
  def __init__(self):
    self.workbook = None
    self.worksheet = None
    self.headers = None
    self._auth = None
    self._session = requests.Session()
    self.do_auth = err
    self.name_idx = 0
    

  def set_auth(self, username, password):
    if len(username) > 0:
      self._session.auth = HttpNtlmAuth(username, password)
    else:
      self._session.auth = None

  def download_sharepoint_xl(self, output_dir, selected_folders):
    hdr = dict(self.headers)
    folders = list(map(lambda x: (hdr[x], '/' if x == 'Path' else ''), selected_folders)) # maintain the order

    for row in tqdm(self.worksheet.iter_rows(min_row=2), total=self.worksheet.max_row-1):
      dirs = map(lambda x: get_valid_filename(row[x[0]].value, x[1]), folders)
      cell = row[self.name_idx]
      self.download_file(cell.hyperlink.target, os.path.join(output_dir, *dirs), cell.value)

  def open_xl(self, file_path):
    self.workbook = openpyxl.load_workbook(file_path, read_only=False) # Need RW to be able to read hyperlinks
    return self.workbook.sheetnames

  def select_ws(self, worksheet_name):
    if (worksheet_name is not None):
      self.worksheet = self.workbook[worksheet_name]
    else:
      self.worksheet = self.workbook.active
    
    self.headers = self.list_headers()
    self.name_idx = next((x[1] for x in self.headers if x[0] == "Name"), 0)

  def list_headers(self):
    return list(map(lambda x: (str(x.value), x.column - 1), self.worksheet[1]))

  def download_file(self, url, dest, name):
    r = self._session.get(url, allow_redirects=True)

    if r.status_code == 401:
      self.do_auth(self._session)
      self.download_file(url, dest, name)
    else:
      os.makedirs(dest, exist_ok=True)
      open(os.path.join(dest, name), 'wb').write(r.content)



def err(session): 
  raise Exception("Unable to access file")

# Adapted from django/utils/text.py
def get_valid_filename(s, extra_chars = ''):
    """
    Return the given string converted to a string that can be used for a clean
    filename. Remove leading and trailing spaces; convert other spaces to
    underscores; and remove anything that is not an alphanumeric, dash,
    underscore, or dot.
    >>> get_valid_filename("john's portrait in 2004.jpg")
    'johns_portrait_in_2004.jpg'
    """
    s = str(s).strip()
    exp = r'(?u)[^- \w.' + extra_chars + ']'
    return re.sub(exp, '', s)
